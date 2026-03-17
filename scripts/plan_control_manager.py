#!/usr/bin/env python3
"""
九星智囊团计划管控系统 V1.0
实现学习计划、训练计划的跟踪管控
"""

import os
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class PlanControlManager:
    """计划管控系统"""

    def __init__(self):
        self.name = "九星智囊团计划管控系统"
        self.version = "1.0.0"
        self.workspace = Path("/root/.openclaw/workspace/03_输出成果")
        self.workspace.mkdir(exist_ok=True)

        # 计划类型
        self.plan_types = {
            "学习计划": {
                "start_date": "2026-03-17",
                "end_date": "2026-04-13",
                "total_days": 28,
                "total_weeks": 4
            },
            "训练计划": {
                "start_date": "2026-03-17",
                "end_date": "2026-04-13",
                "total_days": 28,
                "total_weeks": 4
            }
        }

        # 周目标
        self.weekly_goals = {
            1: {
                "学习": "知识体系搭建，完成7篇知识文档",
                "训练": "团队知识武装，完成简单任务演练"
            },
            2: {
                "学习": "技能工具开发，完成智能调度优化",
                "训练": "协作演练，完成中等复杂度任务"
            },
            3: {
                "学习": "实战验证，完成5个真实任务",
                "训练": "全流程演练，完成复杂项目"
            },
            4: {
                "学习": "持续优化，完成文档体系",
                "训练": "综合演练，能力固化"
            }
        }

        # 成功指标
        self.success_metrics = {
            "学习计划": [
                {"指标": "意图识别准确率", "基线": "85%", "目标": "95%", "权重": "25%"},
                {"指标": "Agent匹配准确率", "基线": "80%", "目标": "90%", "权重": "25%"},
                {"指标": "风险预警覆盖率", "基线": "60%", "目标": "90%", "权重": "20%"},
                {"指标": "知识文档数量", "基线": "1篇", "目标": "20篇", "权重": "15%"},
                {"指标": "实战任务完成数", "基线": "0个", "目标": "5个", "权重": "15%"}
            ],
            "训练计划": [
                {"指标": "平均能力评分", "基线": "90分", "目标": "95分", "权重": "30%"},
                {"指标": "A+级Agent数量", "基线": "0个", "目标": "9个", "权重": "25%"},
                {"指标": "协作效率提升", "基线": "基线", "目标": "+30%", "权重": "20%"},
                {"指标": "任务完成率", "基线": "90%", "目标": "98%", "权重": "15%"},
                {"指标": "训练任务完成数", "基线": "0个", "目标": "27个", "权重": "10%"}
            ]
        }

        # 风险清单
        self.risks = [
            {"风险": "学习时间不足", "可能性": "中", "影响": "高", "应对": "调整学习节奏，优先核心知识"},
            {"风险": "训练任务延期", "可能性": "中", "影响": "中", "应对": "设置缓冲时间，灵活调整"},
            {"风险": "知识质量不达标", "可能性": "低", "影响": "高", "应对": "加强审核机制"},
            {"风险": "团队协作不畅", "可能性": "低", "影响": "高", "应对": "定期协调会议"},
            {"风险": "系统功能不稳定", "可能性": "中", "影响": "高", "应对": "加强测试，快速迭代"}
        ]

    def generate_control_excel(self) -> str:
        """生成计划管控Excel"""
        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="C93832", end_color="C93832", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ===== Sheet 1: 计划概览 =====
        ws1 = wb.active
        ws1.title = "计划概览"

        # 标题
        ws1.merge_cells('A1:G1')
        ws1['A1'] = "九星智囊团计划管控表"
        ws1['A1'].font = Font(bold=True, size=16)
        ws1['A1'].alignment = Alignment(horizontal='center')

        ws1.merge_cells('A2:G2')
        ws1['A2'] = f"制定时间: 2026-03-16 | 执行周期: 2026-03-17 ~ 2026-04-13 | 总计: 4周"
        ws1['A2'].alignment = Alignment(horizontal='center')

        # 计划概览表
        headers = ['计划类型', '开始日期', '结束日期', '总天数', '总周数', '当前状态', '负责人']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        data = [
            ['学习计划', '2026-03-17', '2026-04-13', '28', '4', '🟡 未开始', '🌿 南乔'],
            ['训练计划', '2026-03-17', '2026-04-13', '28', '4', '🟡 未开始', '🌿 南乔']
        ]

        for row_idx, row_data in enumerate(data, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

        # 调整列宽
        for col in range(1, 8):
            ws1.column_dimensions[get_column_letter(col)].width = 15

        # ===== Sheet 2: 周目标 =====
        ws2 = wb.create_sheet("周目标")

        headers2 = ['周次', '时间范围', '学习目标', '训练目标', '关键里程碑', '状态']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        week_data = [
            ['第1周', '03-17 ~ 03-23', '知识体系搭建\n完成7篇知识文档', '团队知识武装\n完成简单任务演练', '知识库建立', '🟡 未开始'],
            ['第2周', '03-24 ~ 03-30', '技能工具开发\n完成智能调度优化', '协作演练\n完成中等复杂度任务', '系统上线', '🟡 未开始'],
            ['第3周', '03-31 ~ 04-06', '实战验证\n完成5个真实任务', '全流程演练\n完成复杂项目', '实战验证', '🟡 未开始'],
            ['第4周', '04-07 ~ 04-13', '持续优化\n完成文档体系', '综合演练\n能力固化', '能力固化', '🟡 未开始']
        ]

        for row_idx, row_data in enumerate(week_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

        # 调整列宽和行高
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 25
        ws2.column_dimensions['D'].width = 25
        ws2.column_dimensions['E'].width = 15
        ws2.column_dimensions['F'].width = 12

        for row in range(2, 6):
            ws2.row_dimensions[row].height = 40

        # ===== Sheet 3: 成功指标 =====
        ws3 = wb.create_sheet("成功指标")

        headers3 = ['计划类型', '指标', '基线', '目标', '权重', '当前值', '完成率', '状态']
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        row_idx = 2
        for plan_type, metrics in self.success_metrics.items():
            for metric in metrics:
                data = [
                    plan_type,
                    metric['指标'],
                    metric['基线'],
                    metric['目标'],
                    metric['权重'],
                    '-',
                    '-',
                    '🟡 未开始'
                ]
                for col_idx, value in enumerate(data, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                row_idx += 1

        # 调整列宽
        for col in range(1, 9):
            ws3.column_dimensions[get_column_letter(col)].width = 15

        # ===== Sheet 4: 风险清单 =====
        ws4 = wb.create_sheet("风险清单")

        headers4 = ['序号', '风险描述', '可能性', '影响程度', '风险等级', '应对措施', '状态']
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for idx, risk in enumerate(self.risks, 1):
            # 计算风险等级
            risk_level = self._calculate_risk_level(risk['可能性'], risk['影响'])

            data = [
                idx,
                risk['风险'],
                risk['可能性'],
                risk['影响'],
                risk_level,
                risk['应对'],
                '🟡 监控中'
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws4.cell(row=idx + 1, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = border

        # 调整列宽
        ws4.column_dimensions['A'].width = 8
        ws4.column_dimensions['B'].width = 20
        ws4.column_dimensions['C'].width = 10
        ws4.column_dimensions['D'].width = 10
        ws4.column_dimensions['E'].width = 12
        ws4.column_dimensions['F'].width = 30
        ws4.column_dimensions['G'].width = 12

        # ===== Sheet 5: 每日跟踪 =====
        ws5 = wb.create_sheet("每日跟踪")

        headers5 = ['日期', '周次', '计划类型', '任务内容', '预计时长', '实际时长', '完成情况', '备注']
        for col, header in enumerate(headers5, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # 生成每日跟踪表（28天）
        start_date = datetime(2026, 3, 17)
        for i in range(28):
            current_date = start_date + timedelta(days=i)
            week_num = (i // 7) + 1
            row_idx = i + 2

            # 上午任务
            ws5.cell(row=row_idx, column=1, value=current_date.strftime('%Y-%m-%d')).border = border
            ws5.cell(row=row_idx, column=2, value=f"第{week_num}周").border = border
            ws5.cell(row=row_idx, column=3, value="学习计划").border = border
            ws5.cell(row=row_idx, column=4, value="知识学习").border = border
            ws5.cell(row=row_idx, column=5, value="2h").border = border
            ws5.cell(row=row_idx, column=6, value="-").border = border
            ws5.cell(row=row_idx, column=7, value="🟡 待完成").border = border
            ws5.cell(row=row_idx, column=8, value="").border = border

        # 调整列宽
        for col in range(1, 9):
            ws5.column_dimensions[get_column_letter(col)].width = 15

        # ===== Sheet 6: 周报汇总 =====
        ws6 = wb.create_sheet("周报汇总")

        headers6 = ['周次', '报告日期', '本周完成情况', '关键成果', '问题与风险', '下周计划', '评分']
        for col, header in enumerate(headers6, 1):
            cell = ws6.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for week in range(1, 5):
            row_idx = week + 1
            ws6.cell(row=row_idx, column=1, value=f"第{week}周").border = border
            ws6.cell(row=row_idx, column=2, value="").border = border
            ws6.cell(row=row_idx, column=3, value="").border = border
            ws6.cell(row=row_idx, column=4, value="").border = border
            ws6.cell(row=row_idx, column=5, value="").border = border
            ws6.cell(row=row_idx, column=6, value="").border = border
            ws6.cell(row=row_idx, column=7, value="-").border = border

        # 调整列宽
        for col in range(1, 8):
            ws6.column_dimensions[get_column_letter(col)].width = 20

        # 保存
        output_path = self.workspace / "九星智囊团计划管控表.xlsx"
        wb.save(output_path)

        return str(output_path)

    def _calculate_risk_level(self, probability: str, impact: str) -> str:
        """计算风险等级"""
        p_map = {"高": 3, "中": 2, "低": 1}
        i_map = {"高": 3, "中": 2, "低": 1}

        score = p_map.get(probability, 1) * i_map.get(impact, 1)

        if score >= 6:
            return "🔴 高"
        elif score >= 3:
            return "🟡 中"
        else:
            return "🟢 低"

    def generate_plan_summary(self) -> str:
        """生成计划管控摘要"""
        summary = f"""
# 九星智囊团计划管控摘要

**生成时间**: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}
**管控周期**: 2026-03-17 ~ 2026-04-13（4周）

---

## 一、计划概览

| 计划类型 | 周期 | 目标 | 状态 |
|----------|------|------|:----:|
| 学习计划 | 4周 | 成为顶级军师 | 🟡 未开始 |
| 训练计划 | 4周 | 打造专家团队 | 🟡 未开始 |

---

## 二、周目标

| 周次 | 学习目标 | 训练目标 |
|:----:|----------|----------|
| 第1周 | 知识体系搭建（7篇文档） | 知识武装 + 简单演练 |
| 第2周 | 技能工具开发（调度优化） | 协作演练 + 中等任务 |
| 第3周 | 实战验证（5个任务） | 全流程演练 + 复杂项目 |
| 第4周 | 持续优化（文档体系） | 综合演练 + 能力固化 |

---

## 三、成功指标

| 指标 | 基线 | 目标 | 提升 |
|------|:----:|:----:|:----:|
| 意图识别准确率 | 85% | 95% | +10% |
| Agent匹配准确率 | 80% | 90% | +10% |
| 风险预警覆盖率 | 60% | 90% | +30% |
| 知识文档数量 | 1篇 | 20篇 | +19篇 |
| 平均能力评分 | 90分 | 95分 | +5分 |
| A+级Agent数量 | 0个 | 9个 | +9个 |

---

## 四、风险管控

| 风险 | 可能性 | 影响 | 应对措施 |
|------|:------:|:----:|----------|
| 学习时间不足 | 中 | 高 | 调整节奏，优先核心 |
| 训练任务延期 | 中 | 中 | 设置缓冲，灵活调整 |
| 知识质量不达标 | 低 | 高 | 加强审核机制 |
| 团队协作不畅 | 低 | 高 | 定期协调会议 |
| 系统功能不稳定 | 中 | 高 | 加强测试，快速迭代 |

---

## 五、管控机制

| 机制 | 频率 | 内容 |
|------|:----:|------|
| 每日跟踪 | 每天 | 任务执行情况 |
| 周报汇总 | 每周 | 完成情况、问题、计划 |
| 月度评估 | 月末 | 目标达成情况、能力评估 |

---

## 六、关键文件

| 文件 | 路径 |
|------|------|
| 计划管控表 | 03_输出成果/九星智囊团计划管控表.xlsx |
| 学习计划 | 学习计划/南乔进化学习计划.md |
| 训练计划 | 学习计划/九星智囊团训练计划.md |
| 每日日报 | 学习日志/YYYY-MM-DD.md |

---

*🌿 南乔 | 九星智囊团*
"""
        return summary


# 演示
if __name__ == "__main__":
    manager = PlanControlManager()

    print("=" * 60)
    print("🎯 九星智囊团计划管控系统 V1.0")
    print("=" * 60)

    # 生成Excel
    print("\n📊 生成计划管控表...")
    excel_path = manager.generate_control_excel()
    print(f"✅ 已生成: {excel_path}")

    # 生成摘要
    print("\n📄 生成计划管控摘要...")
    summary = manager.generate_plan_summary()
    print(summary[:800] + "...")
